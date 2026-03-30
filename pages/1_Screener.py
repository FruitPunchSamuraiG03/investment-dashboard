"""
QUANT SCREENER v8  —  Streamlit Edition
=========================================
Full lineage: Perplexity v4 → CLI v5 → Streamlit v5 → Gemini v7 → v8 (this file)

What's in v8 (everything that survived all iterations + v7 additions):
ENGINE
  • 7-cluster composite score: capital_efficiency / valuation / growth_quality /
    cashflow_quality / dupont_health / balance_sheet / consistency (3-yr ROE/ROIC std dev)
  • Blended momentum multiplier: 1m/3m/6m/12m weighted (12m=40% 6m=30% 3m=20% 1m=10%)
  • Graham Number + discount % scored in valuation cluster
  • EV/FCF ratio with own scoring bands in valuation cluster
  • Piotroski F-Score (9-criteria)
  • Beneish M-Score  (8-variable earnings manipulation flag)
  • Altman Z-Score   (Safe / Grey / Distress zones)
  • Multi-year consistency scoring (3-yr std dev of ROE & ROIC → 1-10)
  • Score percentile (score_pct) alongside raw composite score
HARD GATES
  • Financial-sector bypass: banks/financials exempt from D/E & IntCov gates
  • Missing-data gates: non-financials eliminated if D/E, IntCov, or OPM missing
  • Market cap minimum gate (sidebar slider, default off)
  • Gate thresholds exposed as sidebar sliders
UNIVERSE
  • Local index_files/ folder support (size-range validated, preferred over web)
  • Web fallback with NSE homepage pre-visit + full Chrome session headers
  • Custom watchlist (comma-separated NSE symbols)
  • Nifty 500 / Nifty 750 / Top-30 test
DATA
  • JSON cache with TTL (24h default)
  • Score change tracker: per-universe run history, top improvers/decliners persisted
PORTFOLIO
  • 5 weighting methods: Score-Weighted / Inverse-Vol / Score/Vol / Max-Sharpe / HRP
  • Sector concentration cap (30%), min/max position bounds
  • Benchmark backtest vs Nifty 50 + Nifty 500
UI (7 tabs)
  • Portfolio tab: metric cards, holdings table with gradient
  • Analytics tab: sunburst, factor radar, top-20 bar, 2-yr backtest, correlation heatmap
  • Screener tab: full scored table with multi-timeframe momentum columns
  • Fraud & Distress tab: Beneish caution / Altman Safe-Grey-Distress side-by-side
  • Score Changes tab: improvers/decliners table + delta bar chart
  • Eliminated tab: gate-failed stocks with reasons
  • Failed tab: fetch failures
ALERTS (optional, v7)
  • Email via Gmail SMTP (App Password)
  • WhatsApp via Callmebot (free, no Twilio)
BUG FIXES vs v7
  • Hard gates: removed duplicate unreachable elif branches for financial-sector logic
  • v7 `_cache_path` used `*` instead of `.` in function definition (copy/paste artifact) — fixed
  • All indentation normalised (v7 had mixed tab/space in document)

INSTALL:
    pip install streamlit yfinance openpyxl requests pandas numpy plotly pyportfolioopt scipy
ALERT EXTRAS (optional):
    Email  : Gmail → Settings → Security → App Passwords → generate one
    WhatsApp: https://www.callmebot.com/blog/free-api-whatsapp-messages/
RUN:
    streamlit run quant_screener_v8_streamlit.py
"""

import streamlit as st
import pandas as pd
import numpy as np
import yfinance as yf
import requests
import os, json, warnings, smtplib, ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from io import StringIO, BytesIO
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import plotly.graph_objects as go

warnings.filterwarnings("ignore")

# =============================================================================
# PAGE CONFIG & CSS
# =============================================================================
st.set_page_config(page_title="Quant Screener v8 | TERMINAL", page_icon="⬛", layout="wide")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;600;700&family=Inter:wght@300;400;500;600&display=swap');
    .stApp { background-color: #f4f6f9; color: #1a202c; font-family: 'Inter', sans-serif; }
    .dash-title { font-family: 'JetBrains Mono', monospace; font-size: 1.5rem; font-weight: 700; color: #0057b8; letter-spacing: 0.08em; margin-bottom: 0; }
    .dash-subtitle { font-family: 'JetBrains Mono', monospace; font-size: 0.65rem; color: #64748b; letter-spacing: 0.12em; text-transform: uppercase; margin-bottom: 20px; }
    .section-header { font-family: 'JetBrains Mono', monospace; font-size: 0.6rem; letter-spacing: 0.14em; color: #94a3b8; text-transform: uppercase; border-bottom: 1px solid #e2e8f0; padding-bottom: 6px; margin: 20px 0 10px 0; }
    div[data-testid="stMetric"] { background: #ffffff; border: 1px solid #e2e8f0; border-radius: 8px; padding: 12px 16px; box-shadow: 0 2px 4px rgba(0,0,0,0.03); }
    [data-testid="stMetricLabel"] { font-size: 0.68rem !important; color: #64748b !important; text-transform: uppercase; letter-spacing: 0.08em; }
    [data-testid="stMetricValue"] { font-family: 'JetBrains Mono', monospace !important; font-size: 1.15rem !important; color: #0f172a !important; }
    .stDataFrame { border: 1px solid #e2e8f0; border-radius: 8px; background: #ffffff; box-shadow: 0 2px 4px rgba(0,0,0,0.02); }
    .viz-title { font-family: 'JetBrains Mono', monospace; font-size: 0.85rem; font-weight: 600; color: #475569; margin-bottom: 4px; text-transform: uppercase; letter-spacing: 0.05em; }
</style>
""", unsafe_allow_html=True)

# =============================================================================
# CONFIG
# =============================================================================
CACHE_DIR   = "output/cache/"
OUTPUT_DIR  = "output/"
HISTORY_DIR = "output/"

CONFIG = {
    "max_debt_to_equity":    3.0,
    "min_interest_coverage": 1.5,
    "max_neg_ocf_years":     2,
    "min_market_cap_cr":     0,
    "cluster_weights": {
        "capital_efficiency": 0.26,
        "valuation":          0.23,
        "growth_quality":     0.20,
        "cashflow_quality":   0.12,
        "dupont_health":      0.09,
        "balance_sheet":      0.05,
        "consistency":        0.05,
    },
    "abs_weight":      0.60,
    "rel_weight":      0.40,
    "risk_free_rate":  0.07,
    "price_history":   "2y",
    "max_position":    0.08,
    "min_position":    0.02,
    "max_sector_wt":   0.30,
    "cache_ttl_hours": 24,
}

METHOD_NAMES = {
    1: "Score-Weighted",
    2: "Inverse Volatility",
    3: "Score / Volatility",
    4: "Maximum Sharpe Ratio",
    5: "Hierarchical Risk Parity",
}
METHOD_MAP = {v: k for k, v in METHOD_NAMES.items()}

INDEX_URLS = {
    "nifty500": [
        "https://www.niftyindices.com/IndexConstituent/ind_nifty500list.csv",
        "https://archives.nseindia.com/content/indices/ind_nifty500list.csv",
        "https://raw.githubusercontent.com/kprohith/nse-stock-analysis/master/ind_nifty500list.csv",
    ],
    "microcap250": [
        "https://www.niftyindices.com/IndexConstituent/ind_niftymicrocap250list.csv",
        "https://archives.nseindia.com/content/indices/ind_niftymicrocap250list.csv",
        "https://raw.githubusercontent.com/datasets/nse-indices/main/ind_niftymicrocap250list.csv",
    ],
}

os.makedirs(CACHE_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# =============================================================================
# HELPERS
# =============================================================================
def safe_float(v):
    """Convert any value to float. Returns np.nan for Infinity, NaN, None, non-numeric."""
    if v is None: return np.nan
    try:
        f = float(v)
        return np.nan if (np.isinf(f) or np.isnan(f)) else f
    except: return np.nan

def _get_item(stmt, keys, yr=0):
    if stmt is None: return np.nan
    for k in keys:
        if k in stmt.index:
            vals = stmt.loc[k].dropna().sort_index(ascending=False)
            return safe_float(vals.iloc[yr]) if len(vals) > yr else np.nan
    return np.nan

def fmt(v, suffix="", decimals=2, prefix=""):
    if v is None or (isinstance(v, float) and np.isnan(v)): return "N/A"
    return f"{prefix}{v:.{decimals}f}{suffix}"

# =============================================================================
# CACHE LAYER
# =============================================================================
def _cache_path(symbol):
    return os.path.join(CACHE_DIR, symbol.replace(".", "_") + ".json")

def load_cache(symbol):
    path = _cache_path(symbol)
    if not os.path.exists(path): return None
    try:
        with open(path, "r") as f: c = json.load(f)
        if datetime.now() - datetime.fromisoformat(c["cached_at"]) <= timedelta(hours=CONFIG["cache_ttl_hours"]):
            return c["data"]
    except: pass
    return None

def save_cache(symbol, data):
    try:
        with open(_cache_path(symbol), "w") as f:
            json.dump({"cached_at": datetime.now().isoformat(), "data": data}, f, default=str)
    except: pass

def clear_cache():
    count = 0
    if os.path.exists(CACHE_DIR):
        for f in os.listdir(CACHE_DIR):
            if f.endswith(".json"):
                os.remove(os.path.join(CACHE_DIR, f)); count += 1
    return count

def cache_stats():
    if not os.path.exists(CACHE_DIR): return 0, 0
    files = [f for f in os.listdir(CACHE_DIR) if f.endswith(".json")]
    fresh = 0
    ttl = timedelta(hours=CONFIG["cache_ttl_hours"])
    for f in files:
        try:
            with open(os.path.join(CACHE_DIR, f)) as fp: c = json.load(fp)
            if datetime.now() - datetime.fromisoformat(c["cached_at"]) <= ttl: fresh += 1
        except: pass
    return len(files), fresh

# =============================================================================
# SCORE CHANGE TRACKER
# =============================================================================
def history_path(label): return os.path.join(HISTORY_DIR, f"history_{label}.csv")

def load_history(label):
    p = history_path(label)
    if os.path.exists(p):
        try: return pd.read_csv(p)
        except: pass
    return None

def update_history(df_scored, label):
    cols  = ["symbol","composite_score","rank","piotroski_score","momentum_12m","beneish_m_score"]
    avail = [c for c in cols if c in df_scored.columns]
    out   = df_scored[avail].copy()
    out["run_date"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    out.to_csv(history_path(label), index=False)

def compute_score_changes(df_current, df_history):
    if df_history is None or len(df_history) == 0: return None
    rename = {"composite_score": "prev_score", "rank": "prev_rank",
              "piotroski_score": "prev_piotroski", "momentum_12m": "prev_momentum",
              "run_date": "prev_run_date"}
    hist_cols = ["symbol"] + [c for c in rename if c in df_history.columns]
    merged = df_current[["symbol","composite_score","rank","piotroski_score",
                          "momentum_12m","sector","company_name"]].merge(
        df_history[hist_cols].rename(columns=rename), on="symbol", how="left")
    if "prev_score" in merged.columns:
        merged["score_delta"] = (pd.to_numeric(merged["composite_score"], errors="coerce") -
                                 pd.to_numeric(merged["prev_score"], errors="coerce")).round(2)
    if "prev_rank" in merged.columns:
        merged["rank_delta"] = (pd.to_numeric(merged["prev_rank"], errors="coerce") -
                                pd.to_numeric(merged["rank"], errors="coerce")).apply(
                                lambda x: int(x) if not np.isnan(x) else np.nan)
    return merged

# =============================================================================
# UNIVERSE FETCHER — local files first, web fallback with anti-bot headers
# =============================================================================
@st.cache_data(ttl=3600, show_spinner=False)
def fetch_index_universe(name, custom_symbols=""):
    """Try local index_files/ folder first, fall back to web. Supports custom watchlist."""
    if name == "custom" and custom_symbols:
        syms = [s.strip().upper() for s in custom_symbols.split(",") if s.strip()]
        return pd.DataFrame({"Symbol": syms, "Industry": ["Custom"] * len(syms)})

    symbol_cols = ["Symbol","SYMBOL","symbol","Ticker","TICKER","NSE Symbol","NSE_SYMBOL"]
    size_ranges = {"nifty500": (400, 600), "microcap250": (100, 350)}
    lo, hi = size_ranges.get(name, (50, 9999))

    # ── Step 1: local index_files/ ──
    for base in ["index_files", os.path.join("..", "index_files")]:
        if not os.path.isdir(base): continue
        for fname in sorted(os.listdir(base)):
            if not fname.lower().endswith(".csv"): continue
            try:
                df = pd.read_csv(os.path.join(base, fname))
                df.columns = df.columns.str.strip()
                sym_col = next((c for c in symbol_cols if c in df.columns), None)
                if sym_col is None: continue
                if sym_col != "Symbol": df = df.rename(columns={sym_col: "Symbol"})
                df = df[df["Symbol"].notna() & (df["Symbol"].str.strip() != "")]
                if lo <= len(df) <= hi: return df
            except: continue

    # ── Step 2: web with Chrome session headers + NSE homepage pre-visit ──
    session = requests.Session()
    session.headers.update({
        "User-Agent":                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                                     "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept":                    "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language":           "en-US,en;q=0.5",
        "Connection":                "keep-alive",
        "Upgrade-Insecure-Requests": "1",
    })
    try: session.get("https://www.nseindia.com", timeout=5)
    except: pass

    for url in INDEX_URLS.get(name, []):
        try:
            r = session.get(url, timeout=15)
            if r.status_code == 200:
                df = pd.read_csv(StringIO(r.text))
                df.columns = df.columns.str.strip()
                if "Symbol" not in df.columns: continue
                df = df[df["Symbol"].notna() & (df["Symbol"].str.strip() != "")]
                if lo <= len(df) <= hi: return df
        except: continue

    return pd.DataFrame(columns=["Symbol","Industry"])

# =============================================================================
# PIOTROSKI F-SCORE
# =============================================================================
def compute_piotroski(inc, bs, cf):
    score = 0; details = []
    try:
        def gi(stmt, keys, yr=0):
            if stmt is None: return np.nan
            for k in keys:
                if k in stmt.index:
                    v = stmt.loc[k].dropna().sort_index(ascending=False)
                    return safe_float(v.iloc[yr]) if len(v) > yr else np.nan
            return np.nan
        ni_0  = gi(inc, ["Net Income"]); ni_1  = gi(inc, ["Net Income"], yr=1)
        ocf_0 = gi(cf,  ["Operating Cash Flow","Total Cash From Operating Activities"])
        ta_0  = gi(bs,  ["Total Assets"]); ta_1  = gi(bs,  ["Total Assets"], yr=1)
        ltd_0 = gi(bs,  ["Long Term Debt","Total Debt"]); ltd_1 = gi(bs,  ["Long Term Debt","Total Debt"], yr=1)
        ca_0  = gi(bs,  ["Current Assets","Total Current Assets"]); ca_1 = gi(bs, ["Current Assets","Total Current Assets"], yr=1)
        cl_0  = gi(bs,  ["Current Liabilities","Total Current Liabilities"]); cl_1 = gi(bs, ["Current Liabilities","Total Current Liabilities"], yr=1)
        rev_0 = gi(inc, ["Total Revenue"]); rev_1 = gi(inc, ["Total Revenue"], yr=1)
        cg_0  = gi(inc, ["Cost Of Revenue","Cost of Goods Sold"]); cg_1 = gi(inc, ["Cost Of Revenue","Cost of Goods Sold"], yr=1)
        sh_0  = gi(bs,  ["Ordinary Shares Number","Share Issued","Common Stock"]); sh_1 = gi(bs, ["Ordinary Shares Number","Share Issued","Common Stock"], yr=1)
        roa_0 = ni_0/ta_0 if not np.isnan(ni_0) and not np.isnan(ta_0) and ta_0>0 else np.nan
        roa_1 = ni_1/ta_1 if not np.isnan(ni_1) and not np.isnan(ta_1) and ta_1>0 else np.nan
        def f(cond, lbl):
            nonlocal score; v = 1 if cond else 0; score += v; details.append(f"{lbl}={v}")
        if not np.isnan(roa_0): f(roa_0>0, "F1_ROA+")
        if not np.isnan(ocf_0): f(ocf_0>0, "F2_OCF+")
        if not np.isnan(roa_0) and not np.isnan(roa_1): f(roa_0>roa_1, "F3_ROAimprv")
        if not np.isnan(ocf_0) and not np.isnan(ni_0): f(ocf_0>ni_0, "F4_CFqual")
        if not np.isnan(ltd_0) and not np.isnan(ltd_1) and ta_0>0 and ta_1>0: f(ltd_0/ta_0<ltd_1/ta_1, "F5_Lev-")
        if not np.isnan(ca_0) and not np.isnan(cl_0) and cl_0>0 and cl_1>0: f(ca_0/cl_0>ca_1/cl_1, "F6_CRimprv")
        if not np.isnan(sh_0) and not np.isnan(sh_1): f(sh_0<=sh_1*1.02, "F7_NoDilut")
        if not np.isnan(rev_0) and rev_0>0 and not np.isnan(rev_1) and rev_1>0 and not np.isnan(cg_0) and not np.isnan(cg_1):
            f((rev_0-cg_0)/rev_0>(rev_1-cg_1)/rev_1, "F8_GMimprv")
        if not np.isnan(rev_0) and ta_0>0 and not np.isnan(rev_1) and ta_1>0: f(rev_0/ta_0>rev_1/ta_1, "F9_ATimprv")
    except: pass
    return score, " | ".join(details)

# =============================================================================
# BENEISH M-SCORE
# =============================================================================
def compute_beneish(inc, bs, cf):
    """8-variable Beneish M-Score. M > −2.22 → caution flag (not a hard gate)."""
    try:
        def gi(stmt, keys, yr=0):
            if stmt is None: return np.nan
            for k in keys:
                if k in stmt.index:
                    v = stmt.loc[k].dropna().sort_index(ascending=False)
                    return safe_float(v.iloc[yr]) if len(v) > yr else np.nan
            return np.nan
        rec_0=gi(bs,["Net Receivables","Accounts Receivable"]); rec_1=gi(bs,["Net Receivables","Accounts Receivable"],1)
        rev_0=gi(inc,["Total Revenue"]); rev_1=gi(inc,["Total Revenue"],1)
        cogs_0=gi(inc,["Cost Of Revenue","Cost of Goods Sold"]); cogs_1=gi(inc,["Cost Of Revenue","Cost of Goods Sold"],1)
        ca_0=gi(bs,["Total Current Assets","Current Assets"]); ca_1=gi(bs,["Total Current Assets","Current Assets"],1)
        ppe_0=gi(bs,["Net PPE","Property Plant Equipment Net","Net Property Plant And Equipment"])
        ppe_1=gi(bs,["Net PPE","Property Plant Equipment Net","Net Property Plant And Equipment"],1)
        ta_0=gi(bs,["Total Assets"]); ta_1=gi(bs,["Total Assets"],1)
        dep_0=gi(cf,["Depreciation","Depreciation And Amortization","Depreciation Depletion And Amortization"])
        dep_1=gi(cf,["Depreciation","Depreciation And Amortization","Depreciation Depletion And Amortization"],1)
        sga_0=gi(inc,["Selling General Administrative","Selling General And Administration"])
        sga_1=gi(inc,["Selling General Administrative","Selling General And Administration"],1)
        ltd_0=gi(bs,["Long Term Debt","Total Debt"]); ltd_1=gi(bs,["Long Term Debt","Total Debt"],1)
        cl_0=gi(bs,["Total Current Liabilities","Current Liabilities"]); cl_1=gi(bs,["Total Current Liabilities","Current Liabilities"],1)
        ni_0=gi(inc,["Net Income"]); ocf_0=gi(cf,["Operating Cash Flow","Total Cash From Operating Activities"])
        c = {}
        if all(not np.isnan(x) for x in [rec_0,rev_0,rec_1,rev_1]) and rev_0>0 and rev_1>0:
            c["DSRI"] = (rec_0/rev_0)/(rec_1/rev_1)
        if all(not np.isnan(x) for x in [rev_0,cogs_0,rev_1,cogs_1]) and rev_0>0 and rev_1>0:
            gm0=(rev_0-cogs_0)/rev_0; gm1=(rev_1-cogs_1)/rev_1
            if gm0!=0: c["GMI"] = gm1/gm0
        if all(not np.isnan(x) for x in [ca_0,ppe_0,ta_0,ca_1,ppe_1,ta_1]) and ta_0>0 and ta_1>0:
            a0=1-(ca_0+ppe_0)/ta_0; a1=1-(ca_1+ppe_1)/ta_1
            if a1!=0: c["AQI"] = a0/a1
        if all(not np.isnan(x) for x in [rev_0,rev_1]) and rev_1>0: c["SGI"] = rev_0/rev_1
        if all(not np.isnan(x) for x in [dep_0,ppe_0,dep_1,ppe_1]):
            d0=dep_0/(dep_0+ppe_0) if (dep_0+ppe_0)!=0 else np.nan
            d1=dep_1/(dep_1+ppe_1) if (dep_1+ppe_1)!=0 else np.nan
            if not np.isnan(d0) and not np.isnan(d1) and d0!=0: c["DEPI"] = d1/d0
        if all(not np.isnan(x) for x in [sga_0,rev_0,sga_1,rev_1]) and rev_0>0 and rev_1>0:
            if (sga_1/rev_1)!=0: c["SGAI"] = (sga_0/rev_0)/(sga_1/rev_1)
        if all(not np.isnan(x) for x in [ltd_0,cl_0,ta_0,ltd_1,cl_1,ta_1]) and ta_0>0 and ta_1>0:
            l0=(ltd_0+cl_0)/ta_0; l1=(ltd_1+cl_1)/ta_1
            if l1!=0: c["LVGI"] = l0/l1
        if all(not np.isnan(x) for x in [ni_0,ocf_0,ta_0]) and ta_0>0: c["TATA"] = (ni_0-ocf_0)/ta_0
        if len(c) < 4: return np.nan, "Insufficient data"
        m = -4.84
        for k, coef in {"DSRI":0.920,"GMI":0.528,"AQI":0.404,"SGI":0.892,
                        "DEPI":0.115,"SGAI":-0.172,"LVGI":-0.327,"TATA":4.679}.items():
            if k in c: m += coef * c[k]
        return round(m, 3), "⚠️ CAUTION" if m > -2.22 else "✅ Clean"
    except:
        return np.nan, "Error"

# =============================================================================
# ALTMAN Z-SCORE
# =============================================================================
def compute_altman_z(inc, bs, info):
    """
    Modified Altman Z-Score.
    Z > 2.99 → 🟢 Safe | 1.81–2.99 → 🟡 Grey Zone | Z < 1.81 → 🔴 Distress
    """
    try:
        ta   = _get_item(bs, ["Total Assets"])
        ca   = _get_item(bs, ["Total Current Assets","Current Assets"])
        cl   = _get_item(bs, ["Total Current Liabilities","Current Liabilities"])
        re   = _get_item(bs, ["Retained Earnings"])
        ebit = _get_item(inc, ["Operating Income","EBIT"])
        rev  = _get_item(inc, ["Total Revenue"])
        tl   = _get_item(bs, ["Total Liabilities Net Minority Interest","Total Liabilities"])
        mktcap = safe_float(info.get("marketCap"))
        if any(np.isnan(x) for x in [ta, ca, cl, ebit, rev]) or ta <= 0:
            return np.nan, "Insufficient data"
        wc   = ca - cl
        re_  = re     if not np.isnan(re)     else 0
        tl_  = tl     if not np.isnan(tl)     else ta
        mc_  = mktcap if not np.isnan(mktcap) else 0
        x1 = wc / ta
        x2 = re_ / ta
        x3 = ebit / ta
        x4 = mc_ / tl_ if tl_ > 0 else 0
        x5 = rev / ta
        z = round(1.2*x1 + 1.4*x2 + 3.3*x3 + 0.6*x4 + 1.0*x5, 3)
        zone = "🟢 Safe" if z > 2.99 else ("🟡 Grey Zone" if z > 1.81 else "🔴 Distress")
        return z, zone
    except:
        return np.nan, "Error"

# =============================================================================
# MULTI-YEAR CONSISTENCY SCORING
# =============================================================================
def compute_consistency(inc, bs, cf):
    """3-yr std dev of ROE & ROIC → consistency score 2–10 (lower std = higher score)."""
    result = {}
    try:
        def gi(stmt, keys, yr=0):
            if stmt is None: return np.nan
            for k in keys:
                if k in stmt.index:
                    v = stmt.loc[k].dropna().sort_index(ascending=False)
                    return safe_float(v.iloc[yr]) if len(v) > yr else np.nan
            return np.nan
        def score_std(std):
            if np.isnan(std): return np.nan
            if std < 3:  return 10
            if std < 6:  return 8
            if std < 10: return 6
            if std < 15: return 4
            return 2
        roe_vals = []; roic_vals = []
        for yr in range(3):
            ni     = gi(inc, ["Net Income"], yr)
            equity = gi(bs,  ["Stockholders Equity","Total Stockholder Equity","Common Stock Equity"], yr)
            ebit   = gi(inc, ["Operating Income","EBIT"], yr)
            debt   = gi(bs,  ["Long Term Debt","Total Debt","Long-Term Debt"], yr)
            cash   = gi(bs,  ["Cash And Cash Equivalents","Cash","Cash And Short Term Investments"], yr)
            if not np.isnan(ni) and not np.isnan(equity) and equity > 0:
                roe_vals.append(ni/equity*100)
            if not np.isnan(ebit) and not np.isnan(equity):
                d = debt if not np.isnan(debt) else 0
                c_ = cash if not np.isnan(cash) else 0
                ic = equity + d - c_
                if ic > 0: roic_vals.append(ebit*0.75/ic*100)
        if len(roe_vals) >= 2:
            std = float(np.std(roe_vals, ddof=1))
            result["roe_std_3yr"] = round(std, 2); result["roe_consistency_score"] = score_std(std)
        if len(roic_vals) >= 2:
            std = float(np.std(roic_vals, ddof=1))
            result["roic_std_3yr"] = round(std, 2); result["roic_consistency_score"] = score_std(std)
    except: pass
    return result

# =============================================================================
# STOCK DATA FETCHER
# =============================================================================
def fetch_stock_data(symbol, nse_sector="Unknown"):
    cached = load_cache(symbol)
    if cached: return cached

    result = {"symbol": symbol, "sector": nse_sector, "fetch_status": "ok", "error_msg": ""}
    ticker_obj = None
    for suffix in [".NS", ".BO"]:
        try:
            t = yf.Ticker(symbol + suffix)
            price = safe_float(t.info.get("regularMarketPrice") or t.info.get("currentPrice"))
            if not np.isnan(price) and price > 0:
                ticker_obj = t; result["yf_symbol"] = symbol + suffix; break
        except: continue

    if ticker_obj is None:
        result["fetch_status"] = "failed"
        result["error_msg"]    = "No valid price: delisted / demerged / not on Yahoo Finance"
        return result

    try:
        info   = ticker_obj.info
        pct    = lambda k: safe_float(info.get(k)) * 100 if not np.isnan(safe_float(info.get(k))) else np.nan
        mktcap = safe_float(info.get("marketCap"))
        eps    = safe_float(info.get("trailingEps"))
        bvps   = safe_float(info.get("bookValue"))
        ev     = safe_float(info.get("enterpriseValue"))

        result.update({
            "company_name":        info.get("longName", symbol),
            "yf_sector":           info.get("sector", "Unknown"),
            "industry":            info.get("industry", "Unknown"),
            "market_cap_cr":       mktcap/1e7 if not np.isnan(mktcap) else np.nan,
            "current_price":       safe_float(info.get("regularMarketPrice") or info.get("currentPrice")),
            "beta":                safe_float(info.get("beta")),
            "pe_ttm":              safe_float(info.get("trailingPE")),
            "pb":                  safe_float(info.get("priceToBook")),
            "ps":                  safe_float(info.get("priceToSalesTrailing12Months")),
            "ev_ebitda":           safe_float(info.get("enterpriseToEbitda")),
            "forward_pe":          safe_float(info.get("forwardPE")),
            "dividend_yield":      pct("dividendYield"),
            "roe":                 pct("returnOnEquity"),
            "roa":                 pct("returnOnAssets"),
            "gross_margin":        pct("grossMargins"),
            "operating_margin":    pct("operatingMargins"),
            "net_margin":          pct("profitMargins"),
            "revenue_growth_yoy":  pct("revenueGrowth"),
            "earnings_growth_yoy": pct("earningsGrowth"),
            "debt_to_equity":      safe_float(info.get("debtToEquity"))/100
                                   if not np.isnan(safe_float(info.get("debtToEquity"))) else np.nan,
            "current_ratio":       safe_float(info.get("currentRatio")),
            "momentum_12m":        pct("52WeekChange"),
        })

        # Graham Number
        cp = result["current_price"]
        if not np.isnan(eps) and not np.isnan(bvps) and eps > 0 and bvps > 0:
            g_num = (22.5 * eps * bvps) ** 0.5
            result["graham_number"]       = round(g_num, 2)
            result["graham_discount_pct"] = round((1 - cp/g_num)*100, 2) if not np.isnan(cp) and cp>0 else np.nan

        # Multi-timeframe momentum (6-month history fetch covers all four windows)
        try:
            hist = ticker_obj.history(period="6mo", auto_adjust=True)
            if len(hist) >= 5:
                close = hist["Close"]
                p_now = safe_float(close.iloc[-1])
                def _ret(n_days):
                    idx = max(0, len(close) - n_days - 1)
                    p_past = safe_float(close.iloc[idx])
                    if not np.isnan(p_now) and not np.isnan(p_past) and p_past > 0:
                        return (p_now/p_past - 1) * 100
                    return np.nan
                result["momentum_1m"] = _ret(21)
                result["momentum_3m"] = _ret(63)
                result["momentum_6m"] = _ret(126)
        except: pass

        try:
            inc = ticker_obj.financials
            bs  = ticker_obj.balance_sheet
            cf  = ticker_obj.cashflow

            # Revenue CAGR
            if inc is not None and "Total Revenue" in inc.index:
                rev = inc.loc["Total Revenue"].dropna().sort_index(ascending=False)
                if len(rev) >= 4:
                    r0=safe_float(rev.iloc[0]); r3=safe_float(rev.iloc[3])
                    if not np.isnan(r0) and not np.isnan(r3) and r3>0:
                        result["revenue_cagr_3yr"] = ((r0/r3)**(1/3)-1)*100
                elif len(rev) >= 2:
                    r0=safe_float(rev.iloc[0]); rn=safe_float(rev.iloc[-1]); n=len(rev)-1
                    if not np.isnan(r0) and not np.isnan(rn) and rn>0:
                        result["revenue_cagr_3yr"] = ((r0/rn)**(1/n)-1)*100

            # Op Income CAGR
            for ek in ["Operating Income","EBIT"]:
                if inc is not None and ek in inc.index:
                    op = inc.loc[ek].dropna().sort_index(ascending=False)
                    if len(op) >= 4:
                        o0=safe_float(op.iloc[0]); o3=safe_float(op.iloc[3])
                        if not np.isnan(o0) and not np.isnan(o3) and o3>0 and o0>0:
                            result["op_income_cagr_3yr"] = ((o0/o3)**(1/3)-1)*100
                    break

            ebit = _get_item(inc, ["Operating Income","EBIT"])
            if bs is not None:
                equity = _get_item(bs, ["Stockholders Equity","Total Stockholder Equity","Common Stock Equity"])
                debt   = _get_item(bs, ["Long Term Debt","Total Debt","Long-Term Debt"])
                cash   = _get_item(bs, ["Cash And Cash Equivalents","Cash","Cash And Short Term Investments"])
                assets = _get_item(bs, ["Total Assets"])
                if not np.isnan(ebit) and not np.isnan(equity):
                    ic = equity + (debt if not np.isnan(debt) else 0) - (cash if not np.isnan(cash) else 0)
                    if ic > 0: result["roic"] = ebit*0.75/ic*100
                if not np.isnan(assets) and assets > 0:
                    rev_l = _get_item(inc, ["Total Revenue"])
                    if not np.isnan(rev_l): result["asset_turnover"] = rev_l/assets
                    if not np.isnan(equity) and equity > 0: result["equity_multiplier"] = assets/equity
                int_exp = _get_item(inc, ["Interest Expense","Interest And Debt Expense"])
                if not np.isnan(int_exp) and abs(int_exp)>0 and not np.isnan(ebit):
                    result["interest_coverage"] = ebit/abs(int_exp)

            if cf is not None:
                ocf   = _get_item(cf, ["Operating Cash Flow","Total Cash From Operating Activities"])
                capex = _get_item(cf, ["Capital Expenditure","Capital Expenditures"])
                ni    = _get_item(inc, ["Net Income"])
                if not np.isnan(ocf):
                    result["operating_cf_cr"] = ocf/1e7
                    if not np.isnan(ni) and ni>0: result["cf_quality"] = ocf/ni
                    if not np.isnan(capex):
                        fcf = ocf + capex
                        result["free_cash_flow_cr"] = fcf/1e7
                        if not np.isnan(mktcap) and mktcap>0: result["fcf_yield"] = fcf/mktcap*100
                        if not np.isnan(ni) and ni>0: result["fcf_conversion"] = fcf/ni*100
                        if not np.isnan(ni) and not np.isnan(assets) and assets>0:
                            result["accruals_ratio"] = (ni-fcf)/assets*100
                        # EV/FCF
                        if not np.isnan(ev) and fcf > 0:
                            result["ev_fcf"] = round(ev/fcf, 2)
                for ok in ["Operating Cash Flow","Total Cash From Operating Activities"]:
                    if ok in cf.index:
                        ocf_v = cf.loc[ok].dropna().sort_index(ascending=False)
                        result["neg_ocf_count_3yr"] = int((ocf_v[:3]<0).sum()); break

            pe=result.get("pe_ttm",np.nan); eg=result.get("earnings_growth_yoy",np.nan)
            if not np.isnan(pe) and not np.isnan(eg) and eg>0: result["peg"] = pe/eg

            result["piotroski_score"],  result["piotroski_detail"] = compute_piotroski(inc, bs, cf)
            result["beneish_m_score"],  result["beneish_flag"]     = compute_beneish(inc, bs, cf)
            result["altman_z_score"],   result["altman_zone"]      = compute_altman_z(inc, bs, info)
            result.update(compute_consistency(inc, bs, cf))

        except Exception as e: result["stmt_error"] = str(e)[:150]
    except Exception as e: result["fetch_status"] = "partial"; result["error_msg"] = str(e)[:150]

    save_cache(symbol, result)
    return result

# =============================================================================
# HARD GATES  — financial bypass + missing-data gates + market cap (BUG FIXED)
# =============================================================================
def apply_hard_gates(row):
    reasons = []
    de  = safe_float(row.get("debt_to_equity"))
    ic  = safe_float(row.get("interest_coverage"))
    neg = row.get("neg_ocf_count_3yr", 0) or 0
    opm = safe_float(row.get("operating_margin"))
    mc  = safe_float(row.get("market_cap_cr", np.nan))
    is_fin = "financial" in str(row.get("sector","")).lower() or "bank" in str(row.get("sector","")).lower()

    # Gate 1: D/E — financials exempt from all D/E logic
    if not is_fin:
        if np.isnan(de):                              reasons.append("Missing D/E")
        elif de > CONFIG["max_debt_to_equity"]:       reasons.append(f"D/E={de:.1f}x>{CONFIG['max_debt_to_equity']:.0f}")
    else:
        if not np.isnan(de) and de > CONFIG["max_debt_to_equity"]:
            reasons.append(f"D/E={de:.1f}x>{CONFIG['max_debt_to_equity']:.0f}")

    # Gate 2: Interest coverage — financials exempt
    if not is_fin:
        if np.isnan(ic):                                    reasons.append("Missing IntCov")
        elif ic < CONFIG["min_interest_coverage"]:          reasons.append(f"IntCov={ic:.1f}x<{CONFIG['min_interest_coverage']:.1f}")
    else:
        if not np.isnan(ic) and ic < CONFIG["min_interest_coverage"]:
            reasons.append(f"IntCov={ic:.1f}x<{CONFIG['min_interest_coverage']:.1f}")

    # Gate 3: Negative operating cash flows (applies to all)
    if neg >= CONFIG["max_neg_ocf_years"]: reasons.append(f"NegOCF {int(neg)}/3yrs")

    # Gate 4: Operating margin (applies to all)
    if np.isnan(opm):  reasons.append("Missing OPM")
    elif opm < 0:      reasons.append(f"NegOPM={opm:.1f}%")

    # Gate 5: Market cap minimum (optional, default 0 = disabled)
    min_mc = CONFIG.get("min_market_cap_cr", 0)
    if min_mc > 0 and (np.isnan(mc) or mc < min_mc):
        reasons.append(f"MktCap<{min_mc:.0f}Cr")

    return (False, "; ".join(reasons)) if reasons else (True, "")

# =============================================================================
# SCORING ENGINE
# =============================================================================
VALUATION_BANDS = {
    "pe_ttm":    [(10,10),(15,8),(20,6),(25,4),(30,2),(1e9,0)],
    "pb":        [(1,10),(2,8),(3,6),(4,4),(5,2),(1e9,0)],
    "ps":        [(0.5,10),(1,8),(2,6),(3,4),(5,2),(1e9,0)],
    "ev_ebitda": [(6,10),(9,8),(12,6),(15,4),(20,2),(1e9,0)],
    "peg":       [(0.5,10),(1,8),(1.5,6),(2,4),(3,2),(1e9,0)],
    "ev_fcf":    [(10,10),(15,8),(20,6),(30,4),(40,2),(1e9,0)],
}

def score_tiered(v, bands):
    v = safe_float(v)
    if np.isnan(v): return np.nan
    for upper, sc in bands:
        if v <= upper: return sc
    return 0

def score_hi(v, lo, mid, hi):
    v = safe_float(v)
    if np.isnan(v): return np.nan
    return 10 if v>=hi else (7 if v>=mid else (4 if v>=lo else 1))

def score_lo(v, good, mid, bad):
    v = safe_float(v)
    if np.isnan(v): return np.nan
    return 10 if v<=good else (6 if v<=mid else (3 if v<=bad else 0))

def _blend(a, r):
    a=safe_float(a); r=safe_float(r)
    if np.isnan(a) and np.isnan(r): return np.nan
    if np.isnan(r): return a
    if np.isnan(a): return r
    return a*CONFIG["abs_weight"] + r*CONFIG["rel_weight"]

def _nm(lst):
    v = [safe_float(x) for x in lst if not np.isnan(safe_float(x))]
    return float(np.mean(v)) if v else np.nan

def rel_score(sv, mv, direction="lower_better"):
    sv=safe_float(sv); mv=safe_float(mv)
    if np.isnan(sv) or np.isnan(mv) or mv==0: return np.nan
    ratio=sv/mv
    if direction=="lower_better":
        for t, sc in [(0.50,10),(0.70,8),(0.90,6),(1.10,4),(1.30,2)]:
            if ratio<=t: return sc
        return 0
    else:
        for t, sc in [(2.00,10),(1.50,8),(1.20,6),(1.00,4),(0.80,2)]:
            if ratio>=t: return sc
        return 0

def momentum_multiplier(pct_12m, pct_6m=np.nan, pct_3m=np.nan, pct_1m=np.nan):
    """
    Blended momentum multiplier.
    Weights: 12m=40%, 6m=30%, 3m=20%, 1m=10% — falls back gracefully if shorter windows missing.
    """
    def _mult(pct):
        p = safe_float(pct)
        if np.isnan(p): return np.nan
        if p >= 30: return 1.12
        if p >= 15: return 1.08
        if p >=  5: return 1.04
        if p >=  0: return 1.00
        if p >= -10: return 0.95
        return 0.88
    vals = [(_mult(pct_12m), 0.40), (_mult(pct_6m), 0.30),
            (_mult(pct_3m),  0.20), (_mult(pct_1m), 0.10)]
    avail = [(v, w) for v, w in vals if not np.isnan(v)]
    if not avail: return 1.0
    total_w = sum(w for _, w in avail)
    return sum(v*w for v, w in avail) / total_w

def compute_sector_medians(df):
    metrics = ["pe_ttm","pb","ps","ev_ebitda","roe","roic","roa","fcf_yield"]
    result = {}
    for sector, grp in df.groupby("sector"):
        result[sector] = {}
        for m in metrics:
            if m in grp.columns:
                result[sector][m] = safe_float(pd.to_numeric(grp[m], errors="coerce").median(skipna=True))
    return result

def compute_scores(row, sm_dict):
    s  = {}
    sm = sm_dict.get(row.get("sector","Unknown"), {})
    ra = lambda m: safe_float(row.get(m))
    rm = lambda m: safe_float(sm.get(m))

    s["capital_efficiency"] = _nm([
        _blend(score_hi(ra("roe"),  12,18,25), rel_score(ra("roe"),  rm("roe"),  "higher_better")),
        _blend(score_hi(ra("roic"), 12,18,25), rel_score(ra("roic"), rm("roic"), "higher_better")),
        score_hi(ra("roa"), 5,10,15),
    ])
    s["valuation"] = _nm([
        _blend(score_tiered(ra("pe_ttm"),  VALUATION_BANDS["pe_ttm"]),  rel_score(ra("pe_ttm"),  rm("pe_ttm"))),
        _blend(score_tiered(ra("pb"),      VALUATION_BANDS["pb"]),      rel_score(ra("pb"),      rm("pb"))),
        _blend(score_tiered(ra("ps"),      VALUATION_BANDS["ps"]),      rel_score(ra("ps"),      rm("ps"))),
        score_tiered(ra("ev_ebitda"),  VALUATION_BANDS["ev_ebitda"]),
        score_tiered(ra("peg"),        VALUATION_BANDS["peg"]),
        score_hi(ra("fcf_yield"),      2, 5, 8),
        score_tiered(ra("ev_fcf"),     VALUATION_BANDS["ev_fcf"]),
        score_hi(ra("graham_discount_pct"), -10, 0, 20),
    ])
    s["growth_quality"] = _nm([
        score_hi(ra("revenue_cagr_3yr"),   8, 15, 25),
        score_hi(ra("op_income_cagr_3yr"), 10, 18, 30),
        score_hi(ra("operating_margin"),   8, 15, 25),
        score_hi(ra("revenue_growth_yoy"), 5, 12, 20),
    ])
    s["cashflow_quality"] = _nm([
        score_hi(ra("cf_quality"),    0.8, 1.0, 1.2),
        score_lo(ra("accruals_ratio"), -5,   0,   7),
        score_hi(ra("fcf_conversion"), 50,  70,  90),
    ])
    s["dupont_health"] = _nm([
        score_hi(ra("net_margin"),     5, 10, 20),
        score_hi(ra("asset_turnover"), 0.5, 0.8, 1.2),
        score_lo(ra("equity_multiplier"), 1.5, 2.5, 4.0),
    ])
    s["balance_sheet"] = _nm([
        score_hi(ra("current_ratio"),     1.2, 1.8, 2.5),
        score_hi(ra("interest_coverage"), 3,   6,  10),
        score_lo(ra("debt_to_equity"),    0.3, 0.7, 1.5),
    ])
    s["consistency"] = _nm([
        safe_float(ra("roe_consistency_score")),
        safe_float(ra("roic_consistency_score")),
    ])

    tw=ws=0
    for cluster, w in CONFIG["cluster_weights"].items():
        v = s.get(cluster, np.nan)
        if not np.isnan(safe_float(v)): ws+=v*w; tw+=w
    base = (ws/tw)*10 if tw>0 else np.nan

    mult = momentum_multiplier(
        pct_12m=ra("momentum_12m"), pct_6m=ra("momentum_6m"),
        pct_3m=ra("momentum_3m"),   pct_1m=ra("momentum_1m"),
    )
    s["momentum_multiplier"] = mult
    s["composite_score"] = round(min(100, base*mult), 2) if not np.isnan(safe_float(base)) else np.nan
    return s

# =============================================================================
# ALERT ENGINE — Email (Gmail SMTP) + WhatsApp (Callmebot)
# =============================================================================
def _build_alert_digest(df_scored, score_changes_df, n_top=5):
    lines_txt = []; lines_html = []
    ts = datetime.now().strftime("%d %b %Y %H:%M")
    lines_txt.append(f"QUANT SCREENER v8 — Alert Digest\n{ts}\n{'─'*40}")
    lines_html.append(f"<h2>⬛ Quant Screener v8 — Alert Digest</h2><p><b>{ts}</b></p><hr>")

    top_n = df_scored.head(n_top)[["symbol","company_name","composite_score","pe_ttm","roe","momentum_12m"]]
    lines_txt.append(f"\nTOP {n_top} STOCKS")
    lines_html.append(f"<h3>Top {n_top} Stocks</h3><table border='1' cellpadding='4' style='border-collapse:collapse'>")
    lines_html.append("<tr><th>Symbol</th><th>Name</th><th>Score</th><th>P/E</th><th>ROE%</th><th>Mom12m%</th></tr>")
    for _, r in top_n.iterrows():
        lines_txt.append(f"  {r['symbol']:10s} score={r['composite_score']:.1f} PE={fmt(r.get('pe_ttm'),'',1)} ROE={fmt(r.get('roe'),'%',1)}")
        lines_html.append(f"<tr><td><b>{r['symbol']}</b></td><td>{r.get('company_name','')}</td>"
                          f"<td>{r['composite_score']:.1f}</td><td>{fmt(r.get('pe_ttm'),'',1)}</td>"
                          f"<td>{fmt(r.get('roe'),'%',1)}</td><td>{fmt(r.get('momentum_12m'),'%',1)}</td></tr>")
    lines_html.append("</table>")

    if score_changes_df is not None and "score_delta" in score_changes_df.columns:
        up5 = score_changes_df.dropna(subset=["score_delta"]).nlargest(n_top,"score_delta")
        dn5 = score_changes_df.dropna(subset=["score_delta"]).nsmallest(n_top,"score_delta")
        lines_txt.append(f"\nTOP {n_top} IMPROVERS")
        lines_html.append(f"<h3>Top {n_top} Improvers</h3><table border='1' cellpadding='4' style='border-collapse:collapse'>"
                          "<tr><th>Symbol</th><th>Prev</th><th>Now</th><th>Δ</th></tr>")
        for _, r in up5.iterrows():
            lines_txt.append(f"  {r['symbol']:10s} {fmt(r.get('prev_score'),'',1)} → {r['composite_score']:.1f} (+{r['score_delta']:.2f})")
            lines_html.append(f"<tr><td><b>{r['symbol']}</b></td><td>{fmt(r.get('prev_score'),'',1)}</td>"
                              f"<td>{r['composite_score']:.1f}</td><td style='color:green'>+{r['score_delta']:.2f}</td></tr>")
        lines_html.append("</table>")
        lines_txt.append(f"\nTOP {n_top} DECLINERS")
        lines_html.append(f"<h3>Top {n_top} Decliners</h3><table border='1' cellpadding='4' style='border-collapse:collapse'>"
                          "<tr><th>Symbol</th><th>Prev</th><th>Now</th><th>Δ</th></tr>")
        for _, r in dn5.iterrows():
            lines_txt.append(f"  {r['symbol']:10s} {fmt(r.get('prev_score'),'',1)} → {r['composite_score']:.1f} ({r['score_delta']:.2f})")
            lines_html.append(f"<tr><td><b>{r['symbol']}</b></td><td>{fmt(r.get('prev_score'),'',1)}</td>"
                              f"<td>{r['composite_score']:.1f}</td><td style='color:red'>{r['score_delta']:.2f}</td></tr>")
        lines_html.append("</table>")

    if "beneish_flag" in df_scored.columns:
        caution = df_scored[df_scored["beneish_flag"].str.contains("CAUTION", na=False)][["symbol","composite_score","beneish_m_score"]]
        if len(caution):
            lines_txt.append(f"\nBENEISH CAUTION FLAGS ({len(caution)} stocks)")
            lines_html.append(f"<h3>⚠️ Beneish Caution Flags</h3><table border='1' cellpadding='4' style='border-collapse:collapse'>"
                              "<tr><th>Symbol</th><th>Score</th><th>M-Score</th></tr>")
            for _, r in caution.iterrows():
                lines_txt.append(f"  {r['symbol']:10s} M={fmt(r.get('beneish_m_score'),'',3)}")
                lines_html.append(f"<tr><td><b>{r['symbol']}</b></td><td>{r['composite_score']:.1f}</td>"
                                  f"<td>{fmt(r.get('beneish_m_score'),'',3)}</td></tr>")
            lines_html.append("</table>")

    if "altman_zone" in df_scored.columns:
        distress = df_scored[df_scored["altman_zone"].str.contains("Distress", na=False)][["symbol","composite_score","altman_z_score"]]
        if len(distress):
            lines_txt.append(f"\nALTMAN DISTRESS FLAGS ({len(distress)} stocks)")
            lines_html.append(f"<h3>🔴 Altman Distress Flags</h3><table border='1' cellpadding='4' style='border-collapse:collapse'>"
                              "<tr><th>Symbol</th><th>Score</th><th>Z-Score</th></tr>")
            for _, r in distress.iterrows():
                lines_txt.append(f"  {r['symbol']:10s} Z={fmt(r.get('altman_z_score'),'',3)}")
                lines_html.append(f"<tr><td><b>{r['symbol']}</b></td><td>{r['composite_score']:.1f}</td>"
                                  f"<td>{fmt(r.get('altman_z_score'),'',3)}</td></tr>")
            lines_html.append("</table>")

    lines_txt.append(f"\n{'─'*40}\nGenerated by Quant Screener v8")
    lines_html.append("<hr><p style='color:#94a3b8;font-size:0.8em'>Generated by Quant Screener v8</p>")
    return "\n".join(lines_txt), "".join(lines_html)

def send_email_alert(sender, app_password, recipient, subject, body_html, body_txt):
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject; msg["From"] = sender; msg["To"] = recipient
        msg.attach(MIMEText(body_txt, "plain")); msg.attach(MIMEText(body_html, "html"))
        ctx = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=ctx) as srv:
            srv.login(sender, app_password)
            srv.sendmail(sender, recipient, msg.as_string())
        return True, "Email sent successfully."
    except Exception as e:
        return False, f"Email error: {e}"

def send_whatsapp_alert(phone, apikey, message):
    """Callmebot free WhatsApp API — register at callmebot.com/blog/free-api-whatsapp-messages/"""
    try:
        short_msg = message[:1500] + ("\n…[truncated]" if len(message) > 1500 else "")
        r = requests.get("https://api.callmebot.com/whatsapp.php",
                         params={"phone": phone, "text": short_msg, "apikey": apikey}, timeout=15)
        if r.status_code == 200 and "message queued" in r.text.lower():
            return True, "WhatsApp message queued."
        return False, f"Callmebot ({r.status_code}): {r.text[:120]}"
    except Exception as e:
        return False, f"WhatsApp error: {e}"

# =============================================================================
# PORTFOLIO CONSTRUCTION
# =============================================================================
def build_portfolio(df_scored, n, method, pval):
    if "yf_symbol" not in df_scored.columns: return None, None, None, None
    top_n      = df_scored.dropna(subset=["yf_symbol"]).head(n).copy().reset_index(drop=True)
    yf_symbols = top_n["yf_symbol"].tolist()

    try:
        raw = yf.download(yf_symbols, period=CONFIG["price_history"], auto_adjust=True, progress=False)
        if isinstance(raw.columns, pd.MultiIndex):
            prices = raw.xs("Close", level=0, axis=1) if "Close" in raw.columns.get_level_values(0) else raw
        elif "Close" in raw.columns:
            prices = raw[["Close"]]
        else:
            prices = raw
        if isinstance(prices, pd.Series): prices = prices.to_frame(yf_symbols[0])
        prices  = prices.dropna(axis=1, how="all")
        returns = prices.pct_change().dropna()
    except Exception as e:
        st.error(f"Price Download Error: {e}"); return None, None, None, None

    top_n_avail = top_n[top_n["yf_symbol"].isin(prices.columns)].copy()
    if len(top_n_avail) < 5: return None, None, None, None
    sym_list = top_n_avail["yf_symbol"].tolist()
    vol = returns[sym_list].std() * np.sqrt(252)

    if method == 1:
        sc = top_n_avail.set_index("yf_symbol")["composite_score"].apply(safe_float).fillna(0)
        raw_w = sc / sc.sum()
    elif method == 2:
        iv = 1.0 / vol.replace(0, np.nan).dropna(); raw_w = iv / iv.sum()
    elif method == 3:
        sc = top_n_avail.set_index("yf_symbol")["composite_score"].apply(safe_float).fillna(0)
        sv = sc / (vol * 100 + 1e-9); raw_w = sv / sv.sum()
    elif method == 4:
        try:
            from pypfopt import EfficientFrontier, risk_models, expected_returns
            mu = expected_returns.mean_historical_return(prices[sym_list])
            S  = risk_models.sample_cov(prices[sym_list])
            ef = EfficientFrontier(mu, S, weight_bounds=(CONFIG["min_position"], CONFIG["max_position"]))
            ef.max_sharpe(risk_free_rate=CONFIG["risk_free_rate"])
            raw_w = pd.Series(ef.clean_weights())
        except Exception as e:
            st.warning(f"Max Sharpe failed ({e}). Falling back to Score/Vol.")
            sc = top_n_avail.set_index("yf_symbol")["composite_score"].apply(safe_float).fillna(0)
            sv = sc / (vol * 100 + 1e-9); raw_w = sv / sv.sum()
    elif method == 5:
        try:
            from pypfopt.hierarchical_portfolio import HRPOpt
            hrp = HRPOpt(returns[sym_list]); hrp.optimize()
            raw_w = pd.Series(hrp.clean_weights())
        except Exception as e:
            st.warning(f"HRP failed ({e}). Falling back to Inverse Vol.")
            iv = 1.0 / vol.replace(0, np.nan).dropna(); raw_w = iv / iv.sum()
    else:
        sc = top_n_avail.set_index("yf_symbol")["composite_score"].apply(safe_float).fillna(0)
        sv = sc / (vol * 100 + 1e-9); raw_w = sv / sv.sum()

    raw_w   = raw_w.clip(lower=CONFIG["min_position"], upper=CONFIG["max_position"])
    weights = raw_w / raw_w.sum()

    sym_to_sector = dict(zip(top_n_avail["yf_symbol"], top_n_avail["sector"]))
    sec_wts = {}
    for sym, w in weights.items():
        sec = sym_to_sector.get(sym, "Unknown"); sec_wts[sec] = sec_wts.get(sec, 0) + w
    for sec, sw in list(sec_wts.items()):
        if sw > CONFIG["max_sector_wt"]:
            scale = CONFIG["max_sector_wt"] / sw
            for sym in list(weights.index):
                if sym_to_sector.get(sym) == sec: weights[sym] *= scale
    weights = weights / weights.sum()

    def w_avg(col):
        if col not in top_n_avail.columns: return np.nan
        vals = pd.to_numeric(top_n_avail.set_index("yf_symbol")[col], errors="coerce").reindex(weights.index)
        mask = vals.notna()
        if mask.sum() == 0: return np.nan
        w = weights[mask] / weights[mask].sum()
        return round(float((vals[mask] * w).sum()), 2)

    port_ret = (returns[list(weights.index)] * weights).sum(axis=1)

    # Benchmark comparison
    bench_cumret = pd.DataFrame()
    try:
        bench_raw = yf.download(["^NSEI","^CRSLDX"], period=CONFIG["price_history"], auto_adjust=True, progress=False)
        if isinstance(bench_raw.columns, pd.MultiIndex):
            bp = bench_raw.xs("Close", level=0, axis=1) if "Close" in bench_raw.columns.get_level_values(0) else bench_raw
        elif "Close" in bench_raw.columns:
            bp = bench_raw[["Close"]]
        else:
            bp = bench_raw
        br       = bp.pct_change().dropna()
        common   = port_ret.index.intersection(br.index)
        port_ret = port_ret.loc[common]
        br       = br.loc[common]
        bench_cumret = (1 + br).cumprod()
        bench_cumret.rename(columns={"^NSEI":"Nifty 50","^CRSLDX":"Nifty 500"}, inplace=True, errors="ignore")
    except: pass

    ann_ret = float((1 + port_ret).prod() ** (252/len(port_ret)) - 1)
    ann_vol = float(port_ret.std() * np.sqrt(252))
    cumret  = (1 + port_ret).cumprod()
    mdd     = float(((cumret / cumret.cummax()) - 1).min())
    sharpe  = (ann_ret - CONFIG["risk_free_rate"]) / ann_vol if ann_vol > 0 else np.nan
    hhi     = sum(v**2 for v in sec_wts.values())

    factor_cols = ["capital_efficiency","valuation","growth_quality","cashflow_quality","dupont_health","balance_sheet"]
    factor_scores = {f: safe_float(w_avg(f)) or 0 for f in factor_cols}

    analytics = {
        "ann_ret": ann_ret, "ann_vol": ann_vol, "sharpe": sharpe, "mdd": mdd,
        "beta": w_avg("beta"), "pe": w_avg("pe_ttm"), "pb": w_avg("pb"),
        "roe": w_avg("roe"), "roic": w_avg("roic"), "fcf_yield": w_avg("fcf_yield"),
        "piotroski": w_avg("piotroski_score"), "hhi": hhi, "method": METHOD_NAMES[method],
    }
    chart_data = {
        "cum_returns":   cumret,
        "bench_cumret":  bench_cumret,
        "factor_scores": factor_scores,
        "corr_matrix":   returns[sym_list].corr().fillna(0),
    }

    port_cols = ["symbol","company_name","sector","composite_score","score_pct",
                 "momentum_12m","momentum_6m","momentum_3m","momentum_1m",
                 "piotroski_score","beneish_flag","altman_zone",
                 "pe_ttm","pb","roe","roic","fcf_yield","dividend_yield","beta",
                 "graham_number","graham_discount_pct","ev_fcf",
                 "current_price","market_cap_cr","yf_symbol",
                 "roe_std_3yr","roic_std_3yr",
                 "capital_efficiency","valuation","growth_quality","cashflow_quality"]
    avail_cols = [c for c in port_cols if c in top_n_avail.columns]
    port_df    = top_n_avail[avail_cols].copy()
    port_df["weight_pct"] = port_df["yf_symbol"].map(weights) * 100
    port_df = port_df.sort_values("weight_pct", ascending=False).reset_index(drop=True)
    port_df["rank"] = port_df.index + 1

    if pval:
        port_df["alloc_inr"] = port_df["weight_pct"] / 100 * pval
        def calc_shares(r):
            cp=safe_float(r.get("current_price")); ai=safe_float(r.get("alloc_inr"))
            if np.isnan(cp) or cp<=0 or np.isnan(ai): return np.nan
            return max(1, int(ai/cp))
        port_df["shares_to_buy"] = port_df.apply(calc_shares, axis=1)

    return port_df, analytics, sec_wts, chart_data

# =============================================================================
# EXCEL EXPORT (in-memory buffer)
# =============================================================================
def create_excel_buffer(df_scored, df_gated, failed, port_df, analytics, score_changes_df=None):
    wb   = Workbook()
    navy = PatternFill("solid", fgColor="1F3864"); red = PatternFill("solid", fgColor="7B2C2C")
    grn  = PatternFill("solid", fgColor="1A5C38"); amb = PatternFill("solid", fgColor="7B4B00")
    wh   = Font(color="FFFFFF", bold=True, size=10); ctr = Alignment(horizontal="center")

    def write_df(ws, df, cols, fill):
        avail = [c for c in cols if c in df.columns]
        for ci, col in enumerate(avail, 1):
            cell = ws.cell(row=1, column=ci, value=col.replace("_"," ").title())
            cell.fill=fill; cell.font=wh; cell.alignment=ctr
        for ri, (_, row) in enumerate(df[avail].iterrows(), 2):
            for ci, val in enumerate(row, 1):
                if isinstance(val, (float, np.floating)):
                    v = "" if (pd.isna(val) or np.isinf(val)) else round(float(val), 2)
                elif pd.isna(val): v = ""
                else: v = val
                ws.cell(row=ri, column=ci, value=v).alignment = ctr
        for ci in range(1, len(avail)+1): ws.column_dimensions[get_column_letter(ci)].width = 16
        return avail

    # Sheet 1: Screener Results
    ws1 = wb.active; ws1.title = "Screener Results"
    C1 = ["rank","symbol","company_name","sector","composite_score","score_pct","momentum_multiplier",
          "capital_efficiency","valuation","growth_quality","cashflow_quality","dupont_health","balance_sheet","consistency",
          "piotroski_score","momentum_1m","momentum_3m","momentum_6m","momentum_12m",
          "pe_ttm","pb","ps","ev_ebitda","ev_fcf","peg","dividend_yield",
          "graham_number","graham_discount_pct",
          "roe","roic","roa","fcf_yield","cf_quality","accruals_ratio","fcf_conversion",
          "revenue_cagr_3yr","op_income_cagr_3yr","operating_margin","net_margin",
          "asset_turnover","equity_multiplier","current_ratio","interest_coverage",
          "debt_to_equity","beta","market_cap_cr","current_price",
          "beneish_m_score","beneish_flag","altman_z_score","altman_zone",
          "roe_std_3yr","roic_std_3yr"]
    a1 = write_df(ws1, df_scored, C1, navy)
    if "composite_score" in a1:
        sc_col = get_column_letter(a1.index("composite_score")+1)
        ws1.conditional_formatting.add(f"{sc_col}2:{sc_col}{len(df_scored)+1}",
            ColorScaleRule(start_type="num",start_value=0,  start_color="FF0000",
                           mid_type="num",  mid_value=50,   mid_color="FFFF00",
                           end_type="num",  end_value=100,  end_color="00B050"))
    ws1.freeze_panes = "E2"

    # Sheet 2: Portfolio
    if port_df is not None:
        ws2 = wb.create_sheet("Portfolio")
        C2 = ["rank","symbol","company_name","sector","weight_pct","composite_score","score_pct",
              "momentum_12m","piotroski_score","beneish_flag","altman_zone",
              "pe_ttm","pb","roe","roic","fcf_yield","dividend_yield",
              "graham_number","graham_discount_pct","ev_fcf","beta","current_price",
              "roe_std_3yr","roic_std_3yr"]
        if "alloc_inr" in port_df.columns: C2 += ["alloc_inr","shares_to_buy"]
        write_df(ws2, port_df, C2, grn)
        if analytics:
            sr = len(port_df) + 4
            ws2.cell(row=sr, column=1, value="PORTFOLIO ANALYTICS").font = Font(bold=True, size=12)
            rows = [("Annualised Return",     f"{analytics['ann_ret']*100:.2f}%"),
                    ("Annualised Volatility",  f"{analytics['ann_vol']*100:.2f}%"),
                    ("Sharpe Ratio",           fmt(analytics["sharpe"])),
                    ("Max Drawdown",           f"{analytics['mdd']*100:.2f}%"),
                    ("Weighted Beta",          fmt(analytics["beta"])),
                    ("Weighted P/E",           fmt(analytics["pe"])),
                    ("Weighted ROE (%)",       fmt(analytics["roe"])),
                    ("Weighted ROIC (%)",      fmt(analytics["roic"])),
                    ("Avg Piotroski F-Score",  fmt(analytics["piotroski"])),
                    ("HHI Concentration",      fmt(analytics["hhi"], decimals=4)),
                    ("Weighting Method",       analytics["method"])]
            for i, (k, v) in enumerate(rows, 1):
                ws2.cell(row=sr+i, column=1, value=k).font = Font(bold=True)
                ws2.cell(row=sr+i, column=2, value=v)
            ws2.column_dimensions["A"].width = 35; ws2.column_dimensions["B"].width = 25

    # Sheet 3: Eliminated
    ws3 = wb.create_sheet("Eliminated")
    write_df(ws3, df_gated,
             ["symbol","company_name","sector","gate_reason","debt_to_equity",
              "interest_coverage","neg_ocf_count_3yr","operating_margin","market_cap_cr"], red)

    # Sheet 4: Data Fetch Failed
    ws4 = wb.create_sheet("Data Fetch Failed")
    ws4.cell(row=1,column=1,value="Symbol").font=Font(bold=True)
    ws4.cell(row=1,column=2,value="Likely Reason").font=Font(bold=True)
    for i, item in enumerate(failed, 2):
        ws4.cell(row=i,column=1,value=item.get("symbol",""))
        ws4.cell(row=i,column=2,value=item.get("reason","Delisted / Demerged / Not on Yahoo Finance"))
    ws4.column_dimensions["A"].width=20; ws4.column_dimensions["B"].width=55

    # Sheet 5: Score Changes
    if score_changes_df is not None and "score_delta" in score_changes_df.columns:
        ws5 = wb.create_sheet("Score Changes")
        ws5.cell(row=1,column=1,value="SCORE CHANGE TRACKER").font=Font(bold=True,size=12)
        chg_cols = ["symbol","company_name","sector","rank","prev_rank","rank_delta",
                    "composite_score","prev_score","score_delta","piotroski_score","momentum_12m"]
        avail_cc = [c for c in chg_cols if c in score_changes_df.columns]
        for ci, col in enumerate(avail_cc, 1):
            cell = ws5.cell(row=3, column=ci, value=col.replace("_"," ").title())
            cell.fill=amb; cell.font=wh; cell.alignment=ctr
        top10 = score_changes_df.dropna(subset=["score_delta"]).nlargest(10,"score_delta")
        bot10 = score_changes_df.dropna(subset=["score_delta"]).nsmallest(10,"score_delta")
        ws5.cell(row=2,column=1,value="TOP IMPROVERS").font=Font(bold=True,color="00B050")
        for ri, (_, row) in enumerate(top10[avail_cc].iterrows(), 4):
            for ci, val in enumerate(row, 1):
                if isinstance(val,(float,np.floating)): v="" if (pd.isna(val) or np.isinf(val)) else round(float(val),2)
                elif pd.isna(val): v=""
                else: v=val
                ws5.cell(row=ri,column=ci,value=v).alignment=ctr
        dec_start = 16
        ws5.cell(row=dec_start,column=1,value="TOP DECLINERS").font=Font(bold=True,color="FF0000")
        for ri, (_, row) in enumerate(bot10[avail_cc].iterrows(), dec_start+1):
            for ci, val in enumerate(row, 1):
                if isinstance(val,(float,np.floating)): v="" if (pd.isna(val) or np.isinf(val)) else round(float(val),2)
                elif pd.isna(val): v=""
                else: v=val
                ws5.cell(row=ri,column=ci,value=v).alignment=ctr
        for ci in range(1, len(avail_cc)+1): ws5.column_dimensions[get_column_letter(ci)].width=16
        ws5.freeze_panes="A4"

    buf = BytesIO(); wb.save(buf); return buf.getvalue()

# =============================================================================
# UI — HEADER
# =============================================================================
st.markdown('<div class="dash-title">⬛ QUANT SCREENER v8</div>', unsafe_allow_html=True)
st.markdown('<div class="dash-subtitle">Multi-Factor · Graham · Altman · Beneish · Multi-Timeframe Momentum · EV/FCF · Alerts</div>', unsafe_allow_html=True)
st.markdown("---")

# =============================================================================
# SIDEBAR
# =============================================================================
with st.sidebar:
    st.markdown('<div class="section-header">⚙️ Universe</div>', unsafe_allow_html=True)
    universe_opt = st.selectbox("Index", ["Nifty 500","Nifty 750 (500 + Microcap)","Quick Test (Top 30)","Custom Watchlist"])
    custom_syms  = ""
    if universe_opt == "Custom Watchlist":
        custom_syms = st.text_area("NSE Symbols (comma-separated)", placeholder="TCS,INFY,RELIANCE", height=80)

    st.markdown('<div class="section-header">💼 Portfolio Optimizer</div>', unsafe_allow_html=True)
    n_stocks  = st.slider("Number of Stocks", 10, 40, 20)
    wt_method = st.selectbox("Weighting Method", list(METHOD_NAMES.values()), index=2)
    pval      = st.number_input("Portfolio Value (₹)", min_value=10000, value=500000, step=10000, format="%d")

    st.markdown('<div class="section-header">🛡️ Hard Gate Thresholds</div>', unsafe_allow_html=True)
    CONFIG["max_debt_to_equity"]    = st.slider("Max D/E Ratio",        1.0, 5.0, 3.0, 0.5)
    CONFIG["min_interest_coverage"] = st.slider("Min Interest Coverage", 0.5, 5.0, 1.5, 0.5)
    CONFIG["min_market_cap_cr"]     = st.slider("Min Market Cap (₹ Cr)", 0, 5000, 0, 250)

    st.markdown('<div class="section-header">🗄️ Cache</div>', unsafe_allow_html=True)
    total_c, fresh_c = cache_stats()
    st.caption(f"{total_c} files · {fresh_c} fresh ({CONFIG['cache_ttl_hours']}h TTL)")

    st.markdown('<div class="section-header">🔔 Alerts (Optional)</div>', unsafe_allow_html=True)
    alert_enabled = st.checkbox("Enable alerts after run", value=False)
    alert_email_enabled = alert_wa_enabled = False
    alert_email = alert_app_pw = alert_recipient = alert_wa_phone = alert_wa_apikey = ""
    if alert_enabled:
        alert_email_enabled = st.checkbox("📧 Email (Gmail)")
        if alert_email_enabled:
            alert_email     = st.text_input("Sender Gmail", placeholder="you@gmail.com")
            alert_app_pw    = st.text_input("App Password", type="password",
                                            help="Gmail → Settings → Security → App Passwords")
            alert_recipient = st.text_input("Recipient email")
        alert_wa_enabled = st.checkbox("💬 WhatsApp (Callmebot)")
        if alert_wa_enabled:
            alert_wa_phone  = st.text_input("Phone (intl, no +)", placeholder="919876543210",
                                            help="e.g. 919876543210 for +91 98765 43210")
            alert_wa_apikey = st.text_input("Callmebot API Key", type="password")
        st.caption("Alerts fire once per run after scoring completes.")

    run_btn = st.button("🚀 Run Full Screener", use_container_width=True, type="primary")
    if st.button("🗑️ Clear Cache", use_container_width=True):
        n = clear_cache(); st.toast(f"Cleared {n} cached files."); st.cache_data.clear()

if run_btn:
    st.session_state.update({
        "screener_run":        True,
        "universe_opt":        universe_opt,
        "custom_syms":         custom_syms,
        "screener_label":      ("custom"   if universe_opt == "Custom Watchlist" else
                                "nifty750" if "750" in universe_opt else
                                "test30"   if "Test" in universe_opt else "nifty500"),
        "alert_enabled":       alert_enabled,
        "alert_email_enabled": alert_email_enabled,
        "alert_email":         alert_email,
        "alert_app_pw":        alert_app_pw,
        "alert_recipient":     alert_recipient,
        "alert_wa_enabled":    alert_wa_enabled,
        "alert_wa_phone":      alert_wa_phone,
        "alert_wa_apikey":     alert_wa_apikey,
    })

# =============================================================================
# MAIN RUN
# =============================================================================
if st.session_state.get("screener_run"):
    label    = st.session_state.get("screener_label", "nifty500")
    _uni_opt = st.session_state.get("universe_opt", universe_opt)
    _custom  = st.session_state.get("custom_syms", "")

    status = st.empty()
    prog   = st.progress(0)

    status.info("📥 Fetching index constituents...")
    if label == "custom":
        df_uni = fetch_index_universe("custom", _custom)
    else:
        df_uni = fetch_index_universe("nifty500")
        if "750" in _uni_opt:
            df_micro = fetch_index_universe("microcap250")
            if len(df_micro) > 0:
                df_uni = pd.concat([df_uni, df_micro]).drop_duplicates(subset=["Symbol"])
            else:
                status.warning("⚠️ Microcap 250 unavailable — proceeding with Nifty 500 only.")

    ind_col    = "Industry" if "Industry" in df_uni.columns else df_uni.columns[1]
    symbols    = df_uni["Symbol"].str.strip().tolist()
    sector_map = dict(zip(df_uni["Symbol"].str.strip(), df_uni[ind_col].str.strip()))
    if "Test" in _uni_opt: symbols = symbols[:30]

    records, failed = [], []
    n_total = len(symbols)
    for i, sym in enumerate(symbols):
        status.info(f"🔍 Analysing: **{sym}** ({i+1}/{n_total}) — {n_total-i-1} remaining")
        data = fetch_stock_data(sym, sector_map.get(sym, "Unknown"))
        if data["fetch_status"] != "failed":
            records.append(data)
        else:
            failed.append({"symbol": sym, "reason": data["error_msg"]})
        prog.progress((i+1)/n_total)

    prog.empty()
    if not records:
        status.error("No data fetched. Check symbols or network connection."); st.stop()
    status.success(f"✅ {len(records)} stocks loaded · {len(failed)} failed. Computing scores...")

    df_all = pd.DataFrame(records)
    for col in ["pe_ttm","pb","ps","ev_ebitda","roe","roa","roic","fcf_yield","debt_to_equity",
                "interest_coverage","operating_margin","current_ratio","beta",
                "momentum_12m","revenue_cagr_3yr","op_income_cagr_3yr","market_cap_cr"]:
        if col in df_all.columns: df_all[col] = pd.to_numeric(df_all[col], errors="coerce")
    df_all["sector"] = df_all.apply(
        lambda r: r["sector"] if r["sector"] not in ["Unknown",""] else r.get("yf_sector","Unknown"), axis=1)

    gate_res = df_all.apply(apply_hard_gates, axis=1)
    df_all["passes_gate"] = gate_res.apply(lambda x: x[0])
    df_all["gate_reason"] = gate_res.apply(lambda x: x[1])
    df_passed = df_all[df_all["passes_gate"]].copy().reset_index(drop=True)
    df_gated  = df_all[~df_all["passes_gate"]].copy().reset_index(drop=True)

    if df_passed.empty:
        status.empty()
        st.error("⚠️ Zero stocks passed hard gates. Try relaxing thresholds in the sidebar."); st.stop()

    sector_medians = compute_sector_medians(df_passed)
    score_rows = df_passed.apply(lambda r: compute_scores(r.to_dict(), sector_medians), axis=1)
    df_scored  = pd.concat([df_passed, pd.DataFrame(score_rows.tolist())], axis=1)
    df_scored  = df_scored.sort_values("composite_score", ascending=False).reset_index(drop=True)
    df_scored["rank"]      = df_scored.index + 1
    df_scored["score_pct"] = df_scored["composite_score"].rank(pct=True).mul(100).round(1)

    status.info("⚖️ Optimizing portfolio weights...")
    port_df, analytics, sec_wts, chart_data = build_portfolio(
        df_scored, n_stocks, METHOD_MAP[wt_method], pval)

    df_history       = load_history(label)
    score_changes_df = compute_score_changes(df_scored, df_history)
    update_history(df_scored, label)
    status.empty()

    # ── FIRE ALERTS ──
    if st.session_state.get("alert_enabled"):
        body_txt, body_html = _build_alert_digest(df_scored, score_changes_df)
        subject = f"Quant Screener v8 — {label.upper()} · {datetime.now().strftime('%d %b %Y %H:%M')}"
        if st.session_state.get("alert_email_enabled"):
            ok, msg = send_email_alert(
                st.session_state["alert_email"], st.session_state["alert_app_pw"],
                st.session_state["alert_recipient"], subject, body_html, body_txt)
            (st.success if ok else st.error)(f"📧 Email: {msg}")
        if st.session_state.get("alert_wa_enabled"):
            ok, msg = send_whatsapp_alert(
                st.session_state["alert_wa_phone"], st.session_state["alert_wa_apikey"], body_txt)
            (st.success if ok else st.error)(f"💬 WhatsApp: {msg}")

    # =========================================================================
    # TABS
    # =========================================================================
    tab_port, tab_analytics, tab_screen, tab_fraud, tab_changes, tab_elim, tab_failed = st.tabs([
        "📊 Portfolio", "📈 Analytics", "🏆 Screener",
        "🔬 Fraud & Distress", "📉 Score Changes", "🚫 Eliminated", "⚠️ Failed"
    ])

    # ── PORTFOLIO TAB ──────────────────────────────────────────────────────
    with tab_port:
        if port_df is not None:
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Annual Return",  f"{analytics['ann_ret']*100:.2f}%")
            c2.metric("Volatility",     f"{analytics['ann_vol']*100:.2f}%")
            c3.metric("Sharpe Ratio",   fmt(analytics["sharpe"]))
            c4.metric("Max Drawdown",   f"{analytics['mdd']*100:.2f}%")
            st.markdown("<div style='margin-top:8px'></div>", unsafe_allow_html=True)
            c5, c6, c7, c8 = st.columns(4)
            c5.metric("Weighted Beta",  fmt(analytics["beta"]))
            c6.metric("Weighted P/E",   fmt(analytics["pe"], "x", 1))
            c7.metric("Weighted ROE",   fmt(analytics["roe"], "%", 1))
            c8.metric("Avg Piotroski",  fmt(analytics["piotroski"], "", 1))
            st.markdown("<hr style='margin:18px 0'>", unsafe_allow_html=True)

            st.markdown("<div class='viz-title'>PORTFOLIO HOLDINGS</div>", unsafe_allow_html=True)
            show_cols = ["rank","symbol","company_name","sector","weight_pct"]
            if "shares_to_buy" in port_df.columns: show_cols += ["alloc_inr","shares_to_buy"]
            show_cols += ["composite_score","score_pct","pe_ttm","roe","roic",
                          "piotroski_score","beneish_flag","altman_zone",
                          "graham_discount_pct","ev_fcf"]
            avail_show = [c for c in show_cols if c in port_df.columns]
            fmt_spec = {"weight_pct":"{:.2f}%","composite_score":"{:.1f}","score_pct":"{:.0f}th",
                        "pe_ttm":"{:.1f}","roe":"{:.1f}","roic":"{:.1f}","graham_discount_pct":"{:.1f}%"}
            if "alloc_inr" in avail_show: fmt_spec["alloc_inr"] = "₹{:,.0f}"
            st.dataframe(
                port_df[avail_show].style.format(fmt_spec)
                .background_gradient(subset=["weight_pct"],      cmap="Blues")
                .background_gradient(subset=["composite_score"], cmap="RdYlGn"),
                use_container_width=True, hide_index=True, height=460)

            st.markdown("<br>", unsafe_allow_html=True)
            excel_data = create_excel_buffer(df_scored, df_gated, failed, port_df, analytics, score_changes_df)
            st.download_button(
                "📥 Download Full Excel Report", data=excel_data,
                file_name=f"QuantScreener_v8_{label}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary")
        else:
            st.error("Not enough price data to build portfolio. Try a larger universe.")

    # ── ANALYTICS TAB ──────────────────────────────────────────────────────
    with tab_analytics:
        if port_df is not None and chart_data:
            col_sun, col_alloc = st.columns([1.5, 1], gap="large")
            with col_sun:
                st.markdown("<div class='viz-title'>SECTOR & STOCK DRILL-DOWN</div>", unsafe_allow_html=True)
                st.caption("Click a sector slice to drill into stocks. Click centre to zoom out.")
                labels=["Portfolio"]; parents=[""]; values=[0]; ids=["Portfolio"]
                unique_sectors = port_df["sector"].unique()
                for sec in unique_sectors:
                    ids.append(sec); labels.append(sec); parents.append("Portfolio")
                    values.append(port_df[port_df["sector"]==sec]["weight_pct"].sum())
                for _, r in port_df.iterrows():
                    ids.append(r["symbol"]); labels.append(r["symbol"])
                    parents.append(r["sector"]); values.append(r["weight_pct"])
                values[0] = sum(values[1:len(unique_sectors)+1])
                fig_sun = go.Figure(go.Sunburst(
                    ids=ids, labels=labels, parents=parents, values=values,
                    branchvalues="total", textinfo="label+percent entry",
                    marker=dict(line=dict(color="#ffffff", width=1.5))))
                fig_sun.update_layout(margin=dict(t=10,b=10,l=10,r=10), height=440,
                                      paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(fig_sun, use_container_width=True)
                hhi = analytics["hhi"]
                st.caption(f"HHI: {hhi:.3f}  (0 = fully diversified · 1 = single sector)")
                st.progress(min(hhi, 1.0))

            with col_alloc:
                st.markdown("<div class='viz-title'>SECTOR ALLOCATION</div>", unsafe_allow_html=True)
                sec_df = pd.DataFrame(list(sec_wts.items()), columns=["Sector","Weight"])
                sec_df["Weight"] = (sec_df["Weight"]*100).round(2)
                sec_df = sec_df.sort_values("Weight", ascending=False).reset_index(drop=True)
                st.dataframe(sec_df.style.format({"Weight":"{:.2f}%"})
                             .background_gradient(cmap="Blues"),
                             use_container_width=True, height=440, hide_index=True)

            st.divider()
            col_radar, col_bar = st.columns(2)
            with col_radar:
                st.markdown("<div class='viz-title'>FACTOR EXPOSURE DNA</div>", unsafe_allow_html=True)
                f_labels = [f.replace("_"," ").title() for f in chart_data["factor_scores"]]
                f_vals   = list(chart_data["factor_scores"].values())
                fig_radar = go.Figure(go.Scatterpolar(
                    r=f_vals+[f_vals[0]], theta=f_labels+[f_labels[0]], fill="toself",
                    fillcolor="rgba(14,165,233,0.2)", line=dict(color="#0284c7", width=2)))
                fig_radar.update_layout(
                    polar=dict(radialaxis=dict(visible=True, range=[0,10])),
                    showlegend=False, height=380, margin=dict(l=40,r=40,t=20,b=20),
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(fig_radar, use_container_width=True)

            with col_bar:
                st.markdown("<div class='viz-title'>TOP 20 COMPOSITE SCORES</div>", unsafe_allow_html=True)
                top20 = df_scored.head(20)
                fig_bar = go.Figure(go.Bar(
                    x=top20["composite_score"], y=top20["symbol"], orientation="h",
                    marker_color="#0057b8",
                    text=top20["composite_score"].round(1).astype(str), textposition="outside"))
                fig_bar.update_layout(
                    height=380, margin=dict(t=10,b=10,l=10,r=60),
                    xaxis_title="Composite Score", yaxis=dict(autorange="reversed"),
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(fig_bar, use_container_width=True)

            st.divider()
            st.markdown("<div class='viz-title'>2-YEAR BACKTEST VS BENCHMARK (Base 1.0)</div>", unsafe_allow_html=True)
            fig_bt = go.Figure()
            bench_df = chart_data.get("bench_cumret", pd.DataFrame())
            if not bench_df.empty:
                if "Nifty 500" in bench_df.columns:
                    fig_bt.add_trace(go.Scatter(x=bench_df.index, y=bench_df["Nifty 500"],
                        mode="lines", name="Nifty 500", line=dict(color="#94a3b8", width=1.5, dash="dash")))
                if "Nifty 50" in bench_df.columns:
                    fig_bt.add_trace(go.Scatter(x=bench_df.index, y=bench_df["Nifty 50"],
                        mode="lines", name="Nifty 50", line=dict(color="#d97706", width=1.5, dash="dot")))
            fig_bt.add_trace(go.Scatter(
                x=chart_data["cum_returns"].index, y=chart_data["cum_returns"].values,
                mode="lines", name="Portfolio", fill="tozeroy",
                fillcolor="rgba(0,87,184,0.1)", line=dict(color="#0057b8", width=2.5)))
            fig_bt.update_layout(
                height=400, margin=dict(l=20,r=20,t=20,b=20),
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
            st.plotly_chart(fig_bt, use_container_width=True)

            st.divider()
            st.markdown("<div class='viz-title'>ASSET CORRELATION MATRIX</div>", unsafe_allow_html=True)
            st.caption("−1 = perfectly inverse · +1 = perfectly correlated. Dark red = high co-movement risk.")
            corr = chart_data["corr_matrix"]
            mask = np.triu(np.ones_like(corr, dtype=bool), k=1)
            corr_masked = corr.where(~mask, np.nan)
            fig_corr = go.Figure(go.Heatmap(
                z=corr_masked.values, x=corr.columns, y=corr.index,
                colorscale="RdBu", zmin=-1, zmax=1, hoverongaps=False))
            fig_corr.update_layout(
                yaxis_autorange="reversed", height=580,
                margin=dict(l=20,r=20,t=20,b=20),
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_corr, use_container_width=True)

    # ── SCREENER RESULTS TAB ───────────────────────────────────────────────
    with tab_screen:
        col_stats, col_table = st.columns([1, 4])
        with col_stats:
            st.metric("Passed Gates", len(df_scored))
            st.metric("Eliminated",   len(df_gated))
            st.metric("Failed Fetch", len(failed))
        with col_table:
            screen_cols = ["rank","symbol","company_name","sector","composite_score","score_pct",
                           "piotroski_score","consistency","capital_efficiency","valuation",
                           "growth_quality","cashflow_quality",
                           "pe_ttm","ev_fcf","graham_number","graham_discount_pct",
                           "roe","roic","fcf_yield","dividend_yield",
                           "momentum_1m","momentum_3m","momentum_6m","momentum_12m"]
            avail_sc = [c for c in screen_cols if c in df_scored.columns]
            st.dataframe(
                df_scored[avail_sc].style.format({
                    "composite_score":"{:.1f}","score_pct":"{:.0f}",
                    "consistency":"{:.1f}","capital_efficiency":"{:.1f}",
                    "valuation":"{:.1f}","growth_quality":"{:.1f}","cashflow_quality":"{:.1f}",
                    "pe_ttm":"{:.1f}","ev_fcf":"{:.1f}","graham_number":"₹{:.1f}",
                    "graham_discount_pct":"{:.1f}%","roe":"{:.1f}%","roic":"{:.1f}%",
                    "fcf_yield":"{:.1f}%","dividend_yield":"{:.2f}%",
                    "momentum_1m":"{:.1f}%","momentum_3m":"{:.1f}%",
                    "momentum_6m":"{:.1f}%","momentum_12m":"{:.1f}%",
                }).background_gradient(subset=["composite_score"], cmap="RdYlGn"),
                use_container_width=True, hide_index=True)

    # ── FRAUD & DISTRESS TAB ───────────────────────────────────────────────
    with tab_fraud:
        col_b, col_a = st.columns(2)
        with col_b:
            st.markdown("#### 🔬 Beneish M-Score")
            st.caption("M > −2.22 flags potential earnings manipulation. Not a hard gate.")
            if "beneish_flag" in df_scored.columns:
                caution_df = df_scored[df_scored["beneish_flag"].str.contains("CAUTION", na=False)]
                clean_df   = df_scored[df_scored["beneish_flag"].str.contains("Clean",   na=False)]
                st.markdown(f"⚠️ **{len(caution_df)} flagged** · ✅ **{len(clean_df)} clean**")
                b_cols  = ["rank","symbol","company_name","sector","composite_score","beneish_m_score","beneish_flag","piotroski_score"]
                b_avail = [c for c in b_cols if c in df_scored.columns]
                if len(caution_df):
                    st.markdown("**⚠️ Caution Flags**")
                    st.dataframe(caution_df[b_avail].style.format(
                        {"composite_score":"{:.1f}","beneish_m_score":"{:.3f}"}),
                        use_container_width=True, hide_index=True)
                st.markdown("**✅ Clean — top 30**")
                st.dataframe(clean_df[b_avail].head(30).style.format(
                    {"composite_score":"{:.1f}","beneish_m_score":"{:.3f}"}),
                    use_container_width=True, hide_index=True)
        with col_a:
            st.markdown("#### 📊 Altman Z-Score")
            st.caption("🟢 Safe (Z>2.99) · 🟡 Grey Zone (1.81–2.99) · 🔴 Distress (<1.81)")
            if "altman_zone" in df_scored.columns:
                distress_df = df_scored[df_scored["altman_zone"].str.contains("Distress", na=False)]
                grey_df     = df_scored[df_scored["altman_zone"].str.contains("Grey",     na=False)]
                safe_df     = df_scored[df_scored["altman_zone"].str.contains("Safe",     na=False)]
                st.markdown(f"🔴 **{len(distress_df)} distress** · 🟡 **{len(grey_df)} grey** · 🟢 **{len(safe_df)} safe**")
                a_cols  = ["rank","symbol","company_name","sector","composite_score","altman_z_score","altman_zone","debt_to_equity","current_ratio"]
                a_avail = [c for c in a_cols if c in df_scored.columns]
                if len(distress_df):
                    st.markdown("**🔴 Distress Zone**")
                    st.dataframe(distress_df[a_avail].style.format(
                        {"composite_score":"{:.1f}","altman_z_score":"{:.3f}","debt_to_equity":"{:.2f}"}),
                        use_container_width=True, hide_index=True)
                st.markdown("**🟢 Safe Zone — top 30**")
                st.dataframe(safe_df[a_avail].head(30).style.format(
                    {"composite_score":"{:.1f}","altman_z_score":"{:.3f}"}),
                    use_container_width=True, hide_index=True)

    # ── SCORE CHANGES TAB ─────────────────────────────────────────────────
    with tab_changes:
        if score_changes_df is not None and "score_delta" in score_changes_df.columns:
            prev_date = ""
            if "prev_run_date" in score_changes_df.columns:
                prev_dates = score_changes_df["prev_run_date"].dropna()
                if len(prev_dates) > 0: prev_date = prev_dates.iloc[0]
            st.markdown(f"Comparing to last run: **{prev_date or 'previous run'}**")
            top10 = score_changes_df.dropna(subset=["score_delta"]).nlargest(10,"score_delta")
            bot10 = score_changes_df.dropna(subset=["score_delta"]).nsmallest(10,"score_delta")
            chg_cols  = ["symbol","company_name","sector","rank","prev_rank","rank_delta",
                         "composite_score","prev_score","score_delta"]
            chg_avail = [c for c in chg_cols if c in score_changes_df.columns]
            col_up, col_dn = st.columns(2)
            with col_up:
                st.markdown("#### 🟢 Top Improvers")
                st.dataframe(top10[chg_avail].style.format(
                    {"composite_score":"{:.1f}","prev_score":"{:.1f}","score_delta":"+{:.2f}"}),
                    use_container_width=True, hide_index=True)
            with col_dn:
                st.markdown("#### 🔴 Top Decliners")
                st.dataframe(bot10[chg_avail].style.format(
                    {"composite_score":"{:.1f}","prev_score":"{:.1f}","score_delta":"{:.2f}"}),
                    use_container_width=True, hide_index=True)
            combo  = pd.concat([top10.head(5), bot10.head(5)]).sort_values("score_delta", ascending=False)
            colors = ["#16a34a" if v>=0 else "#dc2626" for v in combo["score_delta"]]
            fig_chg = go.Figure(go.Bar(
                x=combo["symbol"], y=combo["score_delta"], marker_color=colors,
                text=combo["score_delta"].round(2).astype(str), textposition="outside"))
            fig_chg.update_layout(
                title="Score Δ vs Last Run — Top 5 ↑ & ↓", yaxis_title="Score Delta", height=320,
                margin=dict(t=40,b=10), paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_chg, use_container_width=True)
        else:
            st.info(f"No previous run found for **{label}**. History saved — changes appear on next run.")

    # ── ELIMINATED TAB ────────────────────────────────────────────────────
    with tab_elim:
        st.markdown(f"**{len(df_gated)} stocks** eliminated by hard gates.")
        elim_cols = ["symbol","company_name","sector","gate_reason",
                     "debt_to_equity","interest_coverage","neg_ocf_count_3yr",
                     "operating_margin","market_cap_cr"]
        st.dataframe(df_gated[[c for c in elim_cols if c in df_gated.columns]],
                     use_container_width=True, hide_index=True)

    # ── FAILED TAB ────────────────────────────────────────────────────────
    with tab_failed:
        if failed:
            st.markdown(f"**{len(failed)} symbols** could not be fetched.")
            st.dataframe(pd.DataFrame(failed), use_container_width=True, hide_index=True)
        else:
            st.success("✅ All symbols fetched successfully.")
